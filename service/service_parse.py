import openpyxl
import re


def selecting_groups_from_table(url):
    """
    Reads the Excel file, then finds all the groups and writes their coordinates in the form of columns and rows
    to the list.
    :param file url: the path to the Excel table with the schedule
    :return list list_group: Returns a list with the positions of each group (column, row, group name)
    """
    file = openpyxl.load_workbook(url)
    sheet = file.active
    list_group = []
    for column in sheet.iter_cols():
        column_list = []
        for cell in column:
            cell_dict = {}
            reg = r"^(\d(\w|\d)[А-Я])(.*)(\d{3})$"
            if bool(re.search(reg, str(cell.value))):
                cell_dict['value'] = cell.value
                cell_dict['column'] = cell.column
                cell_dict['row'] = cell.row
                column_list.append(cell_dict)
        list_group.append(column_list)
    list_group = [list_column for list_column in list_group if list_column]
    return list_group


def split_group(url):
    """
    Reads Excel table and indexes columns and rows and writes the left and right rows in dictionaries
    :param file url: the path to the Excel table with the schedule
    :return list<dict> more_data: Returns a list of dictionaries that is a row from the table
    """
    file = openpyxl.load_workbook(url)
    sheet = file.active
    more_data = []
    list_group = selecting_groups_from_table(url)
    for column_group in list_group:
        for i in range(len(column_group)):
            row_start_element = column_group[i]['row']
            column = column_group[i]['column']
            row_end_element = column_group[i + 1]['row'] if i != len(column_group) - 1 else row_start_element + 12
            check_empty_lines = 0
            data_lear = []
            index = 0  # Number line.
            for j in range(row_start_element, row_end_element):
                row = {}
                left = sheet.cell(row=j, column=column - 1).value if sheet.cell(row=j,
                                                                                column=column - 1).value is not None else ''
                right = sheet.cell(row=j, column=column).value if sheet.cell(row=j, column=column).value is not None else ''
    
                if left == '' and right == '':
                    check_empty_lines += 1
                    if check_empty_lines == 2:
                        break
                else:
                    check_empty_lines = 0
                row['left'] = left
                row['right'] = right
                data_lear.append(row)
                index += 1
            more_data.append(data_lear)

    return more_data  # [{'left': '', 'right': '9ТМ-281'}, {'left': '', 'right': 'УП 02.01'}...]


def sort_rows_groups(group):
    """ The function converts an unordered representation of Excel table data into a convenient dictionary type.
    Accepts the data type: list.
    Example of conversion:
     [{'left': '', 'right': '9ТМ-281'}, {'left': '', 'right': 'УП 02.01'}...] =>
     {'name_group': '9ТМ-281',
      'lessons': [{'time_lesson': ['08-30'], 'name_lesson': '2п Остермиллер 303', 'more_data': []}...]}
    :param list group: contains an Excel table view by rows
    :return dict: Returns a sorted dictionary that includes the time, the teacher and the name of the discipline
    """
    lesson = {}
    lessons = []
    time_lesson = []
    more_data = []
    check_row_lesson = True
    for row in group[1:]:
        left = row['left']
        right = row['right']

        if left == "" and len(time_lesson) == 0:
            continue

        reg = r"\d{2}-\d{2}"
        if bool(re.search(reg, left)) and check_row_lesson:
            if lesson:
                lesson['more_data'] = more_data
                lessons.append(lesson)
                lesson = {}
                time_lesson = []
                more_data = []

            time_lesson.append(left)
            name_lesson = right
            lesson['time_lesson'] = time_lesson
            lesson['name_lesson'] = right
            if not (group[1]['right'].startswith('УП')):
                check_row_lesson = False
        else:
            if left != '':
                time_lesson.append(left)
            more_data.append(right)
            check_row_lesson = True
    lesson['more_data'] = more_data
    lessons.append(lesson)

    return {
        'name_group': group[0]['right'],
        'lessons': lessons,
    }
